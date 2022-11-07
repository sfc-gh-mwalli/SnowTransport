#imports
import json
import st_connection
import st_connection.snowflake

import streamlit as st
import pandas as pd

from st_aggrid import AgGrid
from st_aggrid.grid_options_builder import GridOptionsBuilder
from st_aggrid import GridUpdateMode, DataReturnMode
from annotated_text import annotated_text
from openpyxl import load_workbook

########## Use wide layout ##########
st.set_page_config(page_icon="🚀", page_title="SnowTransport: Load a file into Snowflake", layout="wide", initial_sidebar_state="expanded")

########## login form and Snowflake connection ##########
session = st.connection.snowflake.login()

########## functions ##########
def get_curr_role():
    crole = pd.DataFrame(session.sql('select current_role() as "name" ').collect())
    crole = crole[["name"]]
    return crole

def get_curr_wh():
    cwh = pd.DataFrame(session.sql('select current_warehouse() as "name" ').collect())
    cwh = cwh[["name"]]
    return cwh

def get_curr_db():
    cdb = pd.DataFrame(session.sql('select current_database() as "name" ').collect())
    cdb = cdb[["name"]]
    return cdb

def get_curr_sc():
    csc = pd.DataFrame(session.sql('select current_schema() as "name" ').collect())
    csc = csc[["name"]]
    return csc

def get_avail_roles():
        query_roles = pd.DataFrame(session.sql('select CURRENT_AVAILABLE_ROLES() as AROLES').collect())
        roles_as_list = json.loads(query_roles.iloc[0,0])
        allroles = pd.DataFrame (roles_as_list, columns = ['name'])
        roles = allroles[["name"]]
        #print(roles)
        return roles
    
def read_sheet(uploaded_file, file_type='xlsx', str_sheetname="", col_list=None, date_col_list=False):
    dataframe = ""
    try:
        if uploaded_file is not None:
            if file_type == 'csv':
                dataframe = pd.read_csv(uploaded_file, usecols=col_list, parse_dates=date_col_list).applymap(
                    lambda s: s.upper() if type(s) == str else s).fillna('')
            else:
                dataframe = pd.read_excel(uploaded_file, sheet_name=str_sheetname, usecols=col_list, parse_dates=date_col_list).applymap(
                    lambda s: s.upper() if type(s) == str else s).fillna('')
        else:
            dataframe = "Empty - Upload a file first"
            st.warning(dataframe)
    except ValueError as e:
        st.error("Problem: " + e.args[0])
    finally:
        return dataframe

def get_sheetnames(upload_file):
    wb = load_workbook(upload_file, read_only=True, keep_links=False)
    return wb.sheetnames
    
########## sidebar ##########
with st.sidebar:
    ########## Role select ##########
        roles_for_select = get_avail_roles()
        current_role = get_curr_role()
        idx = int(roles_for_select[roles_for_select['name']==current_role].index[0])
        #print(roles_for_select)
        role_select = st.selectbox("Role:", 
                                    roles_for_select, 
                                    help="💡 Choose from your available roles. Tip: It is best practice to not use accountadmin",
                                    index=idx
                                    )
        if role_select:
            session.use_role(role_select)
            
        ########## Warehouse select ##########
        warehouses = pd.DataFrame(session.sql('SHOW WAREHOUSES').collect())[['name']]
        if not warehouses.empty:
            select_wh = st.selectbox("Warehouse:",
                                 pd.unique(warehouses['name']),
                                 help='💡 Choices depend on the privileges of selected role',
                                 key='warehouse_select')
            if(select_wh):
                try:
                    session.use_warehouse(select_wh)
                except:
                    st.error('Warehouse privlage error: Please select a different warehouse', icon="🚨")
                    st.stop()
        else:
            select_wh = st.selectbox("Warehouse:",('None Available'),disabled=True)
            st.stop()

        ########## Database Select ##########
        databases = pd.DataFrame(session.sql('SHOW DATABASES').collect())
        #filter out remote databases
        databases = databases[databases['origin'] == '']
        if not databases.empty:
            select_db=st.selectbox("Database:",
                              databases['name'], 
                              help='💡 Choices depend on the privileges of selected role',
                              key='database_select')
            if(select_db):
                try:
                    session.use_database(select_db)
                except:
                    st.warning('Database privlage error: Please select a different database', icon="🚨")
                    st.stop()
        else:
            select_db = st.selectbox("Database:",('None Available'), disabled=True)
            st.warning('This role has no databases available', icon="🚨")
            st.stop()

        ########## Schema select ##########
        schemas = pd.DataFrame(session.sql('SHOW SCHEMAS').collect())
        #filter out information schema
        schemas = schemas[schemas['name'] != 'INFORMATION_SCHEMA']
        schema=st.selectbox("Schema:",
                            schemas['name'], 
                            help='💡 Choices depend on the selected database',
                            key='schema_select')
        if(schema):
            session.use_schema(schema)

        
########## custom animated background style ##########
st.write("""
<style>
[class*="block-container css-18e3th9 egzxvld2"]  {
	background: linear-gradient(-45deg, #ee7752, #e73c7e, #23a6d5, #23d5ab);
	background-size: 500% 500%;
	animation: gradient 15s ease infinite;
	height: 200vh;
}

@keyframes gradient {
	0% {
		background-position: 0% 50%;
	}
	50% {
		background-position: 100% 50%;
	}
	100% {
		background-position: 0% 50%;
	}
}
</style>
""", unsafe_allow_html=True)

########## logo ##########
st.image(
    "rocket2.png",
    width=400
)

########## custom tab css ##########
tabs_font_css = """
<style>
button[data-baseweb="tab"] {
  font-size: 16px;
  border-radius: 5px;
}
</style>
"""

########## display current context ##########
annotated_text(
    (get_curr_role()['name'].iloc[0] ,"current role","#F96815"),
    (get_curr_wh()['name'].iloc[0], "current warehouse","#000C66"),
    (get_curr_db()['name'].iloc[0], "current database","#570861"),
    (get_curr_sc()['name'].iloc[0], "current schema", "#023020")
)
st.info("👈 Set your context via the sidebar.")

########## Tabs ##########
tab1, tab2 = st.tabs(["\u2001 Upload a file \u2001", "\u2001 Explore Tables \u2001"])
st.write(tabs_font_css, unsafe_allow_html=True)

with tab1:
    ########## file upload and table create ##########
    uploaded_file = st.file_uploader(
        "",
        key="1",
        help="Choose a 'csv','xlsx','xls' file",
        type=['csv', 'xlsx', 'xls']
    )

    if uploaded_file is not None:
        if uploaded_file.type == 'text/csv':
            shows = read_sheet(uploaded_file, 'csv')
        else:
            excel_sheets = get_sheetnames(uploaded_file)
            excel_sheets_mod = ("Select Sheet", *excel_sheets)
            sheet = st.selectbox("Select a Workbook Sheet",
                                excel_sheets_mod, help="Select a sheet from this list")
            if sheet != "Select Sheet":
                shows = read_sheet(uploaded_file, 'xls', sheet)
            else:
                st.stop()  # this makes streamlit wait until user selects a worksheet
    else:
        st.info("👆 Upload a csv, xlsx, or xls file from your local system.")
        st.stop()
            
    ########## Setup AG Grid ##########
    gb = GridOptionsBuilder.from_dataframe(shows)
    gb.configure_default_column(enablePivot=True, enableValue=True, enableRowGroup=True, editable=True, groupable=True)
    gb.configure_side_bar()
    gb.configure_pagination(
        enabled=True, paginationAutoPageSize=False, paginationPageSize=20)
    gridOptions = gb.build()

    ########## AG Grid ##########
    response = AgGrid(
        shows,
        gridOptions=gridOptions,
        enable_enterprise_modules=False,
        editable=True,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        fit_columns_on_grid_load=False,
        theme='balham'
    )
    df = pd.DataFrame(response["data"])

    ########## Create Table Form ##########
    with st.form('my_form'):
        st.subheader('Load as a table in Snowflake')
        tablname = st.text_input('Enter the table name')
        submitted = st.form_submit_button('create table')
        if submitted and tablname:
            done = False
            with st.spinner(text="Creating Table in Snowflake..."):
                try:
                    snowpark_df = session.write_pandas(df, tablname.upper(), auto_create_table=True)
                    st.success("TABLE CREATED!", icon="✅")
                    desc = snowpark_df.describe()
                    if desc:
                        with st.spinner(text="Gathering Summary Stats..."):
                            st.dataframe(desc)
                            st.success("Summary Stats complete!", icon="✅")
                            st.snow()
                except:
                    st.error("⚠️ It looks like you don't have privlages to create or overwite a table in this schema. Check your role and try again.")
                
########## tab2 ##########
with tab2:
    st.write("Coming soon...")
          
